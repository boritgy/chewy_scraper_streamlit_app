import streamlit as st
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import time
import json
from stqdm import stqdm
import pandas as pd 
import hmac

st.title("Chewy scraper")

def check_password():

    def password_entered():
        if hmac.compare_digest(st.session_state["password"], st.secrets["password"]):
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Don't store the password.
        else:
            st.session_state["password_correct"] = False

    # Return True if the password is validated.
    if st.session_state.get("password_correct", False):
        return True

    # Show input for password.
    st.text_input(
        "Password", type="password", on_change=password_entered, key="password"
    )
    if "password_correct" in st.session_state:
        st.error("😕 Wrong password")
    return False


if not check_password():
    st.stop()  # Do not continue if check_password is not True.

cookies = {
    'device-id': '501d37e6-79ae-465c-b0c7-9766d38d59dc',
    'pid': '4J3Di1eUQcuZ_RD4u85HmA',
    'sid': '1839bb0a-1971-4312-b667-0c883a662da3',
    'x-feature-preview': 'false',
    'abTestingAnonymousPID': 'K9AjefB4QVy6zeKGnFqADg',
    '_gcl_au': '1.1.1916444144.1726562974',
    '_ga': 'GA1.1.39900230.1726562974',
    'rxVisitor': '172656297441797CIT7PKGJ3VDF1O2ETEFSK4KL1U1URM',
    '_fbp': 'fb.1.1726562974779.68070706348477632',
    '_mibhv': 'anon-1726562974923-2204694577_6593',
    'fpGetInitStatus': 'SuccessfulResponse',
    '_tt_enable_cookie': '1',
    '_ttp': 'dTv3piPSJT1P5m5IYcWHUhi4Juj',
    'ajs_anonymous_id': '0e516ba1-4662-4a70-8628-bff483311dd7',
    'fpPostInitStatus': 'SuccessfulResponse',
    'addshoppers.com': '2%7C1%3A0%7C10%3A1726508880%7C15%3Aaddshoppers.com%7C44%3AMDhiMWQ4MmU4NmI4NDI2NWJhMzhjNzNmMWQxZTYyNGQ%3D%7C80e05a031f510e961d9c521fa68174f4ab6d12b458840624879e4522863c980e',
    'experiment_': '',
    'AKA_A2': 'A',
    'bm_sz': 'EC937A05B71F8A74CD9726A2B3EF308D~YAAQilgWApTdNGWSAQAAItcBaxlMpqMzz5iIgI9CYz+L+Oh+24Yb6T6D3nvgGpB1CqSljuOEimS/yody4qSMBv5ahugsc4maugMDmLYbNr1Wf/7U202CEnLnr9buc0RR1vH3UQAyndqeN5dPkHY7+l83sxWZkZ5aqw86HF2GS+lu60ysCxKauwbxdZUecLtcCJPdHijYxXXedVoFHPe7unGu/gkjfMEHlREF59jW5STQp/FfTdns0gu/XOou6jJ6k3CTIyh9yL7y2bwPJjD66S1cYXfMw72aHfAbM0IKP0dSOz7eTpBvkHeUtsJiMqosom82wz0rVAEICEuRVDLVrPHxWI7p78Cg3GGgDIiguhvDKm899HicIyHtViV7RlOJ8SqcIG/pArjoklgGYg==~3486006~3294277',
    'chewy-insights': 'ts=https://www.chewy.com/',
    'bm_sv': 'E411BE73331DF1B4B7B68DF90F062138~YAAQhlgWApQZCTGSAQAAI9sBaxmXXMtVu9M2XaraaVHQGqDndA2e4KC6xDNfzs5sqodYS3Gj4jWDvmo1L19C+5keZV8czqvPa6V6RSkvBz+omnqIGqRnBaVhC3SpODFHF+RhHLvW6mZ2cvsUTZJ5Mhm92ZdIjBqq1dg1ey6k5H7so5Vypu5kb6OvrzTimyPUlIh5odfrHsawBDH3n5Kb77cL82CM8/jkq2kSByQhKpV9K2Kg3zjUBqYC7pTTiJ0=~1',
    'sitever_practicehubofficeuse_ui': 'old_backend',
    'sitever_chewyportalpdpui': 'new',
    'sitever_chewy_pethealthui_apt_ui': 'old_backend',
    'bm_mi': 'AC3405276601E09FF36DEB933CD9E1F0~YAAQhlgWApwZCTGSAQAAmd0BaxkxhDqv9Uw4soLphJuQYb/9WqaktSVpWYXtY0pCP7/Z2+0FdRIc9JgiE7htzQxC6zFBM6MtUiPvTHLNrRb9VCnofhKw2gqK6cFOwXLR2YwvwtuvRjlQJ02lSKI12hr1W5VcOJw2hjVMN97fJ/nJMxkprxKNqaAeBQCXiSEjzm58O0/HI6IcELdB3AQObTCQm4olWAiCYbL0dcDtorHLaObFtNkmX4o9EWKb2ffixNuXXvMS8eyc575gRajMRyf/E7g5NnAYNsZ75qtt7fsq+lE+kYupU5Bvmrz7~1',
    'ak_bmsc': '3244C3D75140074CF982712F0335FD5E~000000000000000000000000000000~YAAQhlgWAtEZCTGSAQAAA/EBaxlDB/ziGD15EVPdcAwDPbsX2WirMIxGyKOhNSE5bhYWpZzrgcxiczj6hC3Y4Oh6bYcl2UYL5RnabHaqiNk+jHhwkC1f7jzPApqF42MvWJMT8PdjKBTILHkbHSwWsX9u1RimGJFPBXPQRvykbDqGFsZTDQfDGpT5+RPOu/wVA4h4TZLjgo9zzBKugcDYDdaL+MHtzN8m/EgoEvrM+14clqFpkS9xLws8QQvMZ7alVUaWzGN+YeY3ZO1xtwN3dL2CjeeD+yrK0pjoCTNY8HTI//ytXdefsKMVPm+EVSEzoZpFv5nvGTxa6Y3aDAuhA6rzI0EXsdLEUBILGA9ZlpQzTtHdQlLfAQlFaNCLF2a5MD8NlyTIo4RavMT5uOG2VQIjZJID4w2ktbKQGXNUhnNTIVPL6awlfRm+jBqsFrJYWxpX2hrQwYf8m2VFWb+8Du59BNe/cGXxtNY=',
    'dtCookie': 'v_4_srv_12_sn_P0BCVKEJ8PU0IAQ36NHB98TJMEJCQNOV_app-3A7077613abb396c51_1_ol_0_perc_100000_mul_1_rcs-3Acss_0',
    '_clck': '7tp568%7C2%7Cfpu%7C0%7C1720',
    '_iidt': 'HLMN2MOyEIgMPEPxM7Q3o5hpt9tHP3yBdGakb1B+NSjTxmZ6lDRrUEQQujAkx6worFJ9jGeprzTlUA==',
    '_vid_t': 'n0gF8g0y5w3v+wRFCgB9WewPH8hYrWfn3S4xjKBqgT/yCVOe0QU0wl13x7NHXF9cIWd+a1DCBhwCsQ==',
    'fppro_id': '{"rid":"1728372147839.MMnxPJ","vid":"z8RX4Dez2DdvMwIkaMwe","exp":1728976947617}',
    'pageviewCount': '2',
    'dtSa': '-',
    '_uetsid': '11cac770854611efb06f6ff8eb24ccf8',
    '_uetvid': 'c38d6f3074d111efaa4d197904445237',
    'pageviewCount30m': '2',
    'OptanonConsent': 'isGpcEnabled=0&datestamp=Tue+Oct+08+2024+09%3A22%3A39+GMT%2B0200+(k%C3%B6z%C3%A9p-eur%C3%B3pai+ny%C3%A1ri+id%C5%91)&version=202304.1.0&browserGpcFlag=0&isIABGlobal=false&hosts=&landingPath=NotLandingPage&groups=BG36%3A1%2CC0004%3A1%2CC0010%3A1%2CC0011%3A1%2CC0001%3A1%2CC0003%3A1%2CC0002%3A1&AwaitingReconsent=false',
    '_abck': 'D72CA898F39A1BD80917F567A9F1B7E1~-1~YAAQhlgWAsYaCTGSAQAAgjgCawxgPDGSUp47ygYDLudlqcmMF+xb/xnuKe9Hx9XvACuQ0mPaBhktnBWmByq2zKG2BOoy0Tp/Fz6zsl6mi2wJAUHae1wuiV/QHvEhtc9bovkUDgjgV7hGObeD98MOvdV14DbgY+yYapR/czvPsHO/SzB0bo6Dv/EWjFktE+P/les1PBxGmdHkD4yqy0vXbYWzLjMf+VRnvA323ZNzk4A4c6MsQC3PFuDgyP6LVlaJ5Sx7fIIGauydgDIzqePIdS/PLTJKIljV6tzkGzF2M9H4pvSq3nsgMrUKAR09jIbTLdaOxpcburtOqDlVwbHW7i7YngVNGNTvOGrQKtvL93VF+eTMVQUTQqoEGhdWn5JrNryfMdWFdC3vfA9r+aE7+r4SWWzdi/XaDWcXaSjGUveGSRJYUQrNvH2tKV+EGXYKNfXFxg6pK2hXmGsQZSyXwWVA7mzIsbKdkycRfGavWMYvrADWz/KDtwdg5LiuXkyUhgfwZrtAQE50k3Dlg3MLl27TQqc=~0~-1~-1',
    'KP_UIDz-ssn': '02rcgiwlC6QcdD4HSyDzO84ov5sraIHxL9qPlyrTOT4H6k7LyLD9RoKI4IgSwLDs7KAco9TPuQP45bKQSe1Eknb2gfgllUoQfVYhgsDzNpu8iPwnpY4FLfAoCrvQZY7vz74sJ1PpBV6vSDCbCkYkJbYa9PsRjlYQKwGb6Bd7sV',
    'KP_UIDz': '02rcgiwlC6QcdD4HSyDzO84ov5sraIHxL9qPlyrTOT4H6k7LyLD9RoKI4IgSwLDs7KAco9TPuQP45bKQSe1Eknb2gfgllUoQfVYhgsDzNpu8iPwnpY4FLfAoCrvQZY7vz74sJ1PpBV6vSDCbCkYkJbYa9PsRjlYQKwGb6Bd7sV',
    'plpids': '28fm:3qyi|284v:284w|2qlr:3dld|68i3:68i4|39ac:nvzy|286r:286s|2qnk:2ryt|2883:r8wm|3mr5:3mr6|do7i:gvye|2bgh:3tls|6e87:r93i|28h8:28h9|mbfy:q29q|2qlj:3dle|39ah:nvym|2s7y:r93a|gr7q:gr7y|3r12:3r13|k01q:k0gu|3l0l:3l0m|32g4:32g5|28bh:28bj|3mni:3mnj|3mr3:3mr4|28g9:r8xy|28a8:28a9|2ql8:2qla|lt8e:nhx2|28jz:28k0|iud2:iuda|l3ri:l3rq|4cuk:4cul|2qnc:2y4v|l3qm:l3qu|2855:r8yu',
    '_ga_GM4GWYGVKP': 'GS1.1.1728372145.8.1.1728372164.41.0.0',
    '_clsk': '12df93h%7C1728372164793%7C3%7C0%7Cr.clarity.ms%2Fcollect',
    'rxvt': '1728373965004|1728372145522',
    'akavpau_defaultvp': '1728372465~id=b14ffcbfec292c291bbfdce4efcfabc0',
    'akaalb_chewy_ALB': '1728372765~op=prd_chewy_lando:www-prd-lando-use2|prd_chewy_plp:www-prd-plp-use2|prd_kasada_plp:prd-kasada-plp-use2|chewy_com_ALB:www-chewy-use2|kasada_prd:kasada-prd|prd_kasada:prd-kasada-haproxy-use1|~rv=86~m=www-prd-lando-use2:0|www-prd-plp-use2:0|prd-kasada-plp-use2:0|www-chewy-use2:0|kasada-prd:0|prd-kasada-haproxy-use1:0|~os=43a06daff4514d805d02d3b6b5e79808~id=e0256312fba6f14d298097936f1f798f',
    'dtPC': '12$372153273_297h-vCCMFKGHPEBLUOPKIIFDMIPHJKGPMSLIH-0e0',
    'RT': '"z=1&dm=www.chewy.com&si=843cf466-6940-48be-95fd-c6abedbb3823&ss=m20457l9&sl=2&tt=4qz&rl=1&nu=3e3ez51&cl=igw&obo=1&ld=1efx&r=3e3ez51&ul=1efx"',
}

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'hu-HU,hu;q=0.9,el-GR;q=0.8,el;q=0.7,en-US;q=0.6,en;q=0.5',
    'cache-control': 'max-age=0',
    # 'cookie': 'device-id=501d37e6-79ae-465c-b0c7-9766d38d59dc; pid=4J3Di1eUQcuZ_RD4u85HmA; sid=1839bb0a-1971-4312-b667-0c883a662da3; x-feature-preview=false; abTestingAnonymousPID=K9AjefB4QVy6zeKGnFqADg; _gcl_au=1.1.1916444144.1726562974; _ga=GA1.1.39900230.1726562974; rxVisitor=172656297441797CIT7PKGJ3VDF1O2ETEFSK4KL1U1URM; _fbp=fb.1.1726562974779.68070706348477632; _mibhv=anon-1726562974923-2204694577_6593; fpGetInitStatus=SuccessfulResponse; _tt_enable_cookie=1; _ttp=dTv3piPSJT1P5m5IYcWHUhi4Juj; ajs_anonymous_id=0e516ba1-4662-4a70-8628-bff483311dd7; fpPostInitStatus=SuccessfulResponse; addshoppers.com=2%7C1%3A0%7C10%3A1726508880%7C15%3Aaddshoppers.com%7C44%3AMDhiMWQ4MmU4NmI4NDI2NWJhMzhjNzNmMWQxZTYyNGQ%3D%7C80e05a031f510e961d9c521fa68174f4ab6d12b458840624879e4522863c980e; experiment_=; AKA_A2=A; bm_sz=EC937A05B71F8A74CD9726A2B3EF308D~YAAQilgWApTdNGWSAQAAItcBaxlMpqMzz5iIgI9CYz+L+Oh+24Yb6T6D3nvgGpB1CqSljuOEimS/yody4qSMBv5ahugsc4maugMDmLYbNr1Wf/7U202CEnLnr9buc0RR1vH3UQAyndqeN5dPkHY7+l83sxWZkZ5aqw86HF2GS+lu60ysCxKauwbxdZUecLtcCJPdHijYxXXedVoFHPe7unGu/gkjfMEHlREF59jW5STQp/FfTdns0gu/XOou6jJ6k3CTIyh9yL7y2bwPJjD66S1cYXfMw72aHfAbM0IKP0dSOz7eTpBvkHeUtsJiMqosom82wz0rVAEICEuRVDLVrPHxWI7p78Cg3GGgDIiguhvDKm899HicIyHtViV7RlOJ8SqcIG/pArjoklgGYg==~3486006~3294277; chewy-insights=ts=https://www.chewy.com/; bm_sv=E411BE73331DF1B4B7B68DF90F062138~YAAQhlgWApQZCTGSAQAAI9sBaxmXXMtVu9M2XaraaVHQGqDndA2e4KC6xDNfzs5sqodYS3Gj4jWDvmo1L19C+5keZV8czqvPa6V6RSkvBz+omnqIGqRnBaVhC3SpODFHF+RhHLvW6mZ2cvsUTZJ5Mhm92ZdIjBqq1dg1ey6k5H7so5Vypu5kb6OvrzTimyPUlIh5odfrHsawBDH3n5Kb77cL82CM8/jkq2kSByQhKpV9K2Kg3zjUBqYC7pTTiJ0=~1; sitever_practicehubofficeuse_ui=old_backend; sitever_chewyportalpdpui=new; sitever_chewy_pethealthui_apt_ui=old_backend; bm_mi=AC3405276601E09FF36DEB933CD9E1F0~YAAQhlgWApwZCTGSAQAAmd0BaxkxhDqv9Uw4soLphJuQYb/9WqaktSVpWYXtY0pCP7/Z2+0FdRIc9JgiE7htzQxC6zFBM6MtUiPvTHLNrRb9VCnofhKw2gqK6cFOwXLR2YwvwtuvRjlQJ02lSKI12hr1W5VcOJw2hjVMN97fJ/nJMxkprxKNqaAeBQCXiSEjzm58O0/HI6IcELdB3AQObTCQm4olWAiCYbL0dcDtorHLaObFtNkmX4o9EWKb2ffixNuXXvMS8eyc575gRajMRyf/E7g5NnAYNsZ75qtt7fsq+lE+kYupU5Bvmrz7~1; ak_bmsc=3244C3D75140074CF982712F0335FD5E~000000000000000000000000000000~YAAQhlgWAtEZCTGSAQAAA/EBaxlDB/ziGD15EVPdcAwDPbsX2WirMIxGyKOhNSE5bhYWpZzrgcxiczj6hC3Y4Oh6bYcl2UYL5RnabHaqiNk+jHhwkC1f7jzPApqF42MvWJMT8PdjKBTILHkbHSwWsX9u1RimGJFPBXPQRvykbDqGFsZTDQfDGpT5+RPOu/wVA4h4TZLjgo9zzBKugcDYDdaL+MHtzN8m/EgoEvrM+14clqFpkS9xLws8QQvMZ7alVUaWzGN+YeY3ZO1xtwN3dL2CjeeD+yrK0pjoCTNY8HTI//ytXdefsKMVPm+EVSEzoZpFv5nvGTxa6Y3aDAuhA6rzI0EXsdLEUBILGA9ZlpQzTtHdQlLfAQlFaNCLF2a5MD8NlyTIo4RavMT5uOG2VQIjZJID4w2ktbKQGXNUhnNTIVPL6awlfRm+jBqsFrJYWxpX2hrQwYf8m2VFWb+8Du59BNe/cGXxtNY=; dtCookie=v_4_srv_12_sn_P0BCVKEJ8PU0IAQ36NHB98TJMEJCQNOV_app-3A7077613abb396c51_1_ol_0_perc_100000_mul_1_rcs-3Acss_0; _clck=7tp568%7C2%7Cfpu%7C0%7C1720; _iidt=HLMN2MOyEIgMPEPxM7Q3o5hpt9tHP3yBdGakb1B+NSjTxmZ6lDRrUEQQujAkx6worFJ9jGeprzTlUA==; _vid_t=n0gF8g0y5w3v+wRFCgB9WewPH8hYrWfn3S4xjKBqgT/yCVOe0QU0wl13x7NHXF9cIWd+a1DCBhwCsQ==; fppro_id={"rid":"1728372147839.MMnxPJ","vid":"z8RX4Dez2DdvMwIkaMwe","exp":1728976947617}; pageviewCount=2; dtSa=-; _uetsid=11cac770854611efb06f6ff8eb24ccf8; _uetvid=c38d6f3074d111efaa4d197904445237; pageviewCount30m=2; OptanonConsent=isGpcEnabled=0&datestamp=Tue+Oct+08+2024+09%3A22%3A39+GMT%2B0200+(k%C3%B6z%C3%A9p-eur%C3%B3pai+ny%C3%A1ri+id%C5%91)&version=202304.1.0&browserGpcFlag=0&isIABGlobal=false&hosts=&landingPath=NotLandingPage&groups=BG36%3A1%2CC0004%3A1%2CC0010%3A1%2CC0011%3A1%2CC0001%3A1%2CC0003%3A1%2CC0002%3A1&AwaitingReconsent=false; _abck=D72CA898F39A1BD80917F567A9F1B7E1~-1~YAAQhlgWAsYaCTGSAQAAgjgCawxgPDGSUp47ygYDLudlqcmMF+xb/xnuKe9Hx9XvACuQ0mPaBhktnBWmByq2zKG2BOoy0Tp/Fz6zsl6mi2wJAUHae1wuiV/QHvEhtc9bovkUDgjgV7hGObeD98MOvdV14DbgY+yYapR/czvPsHO/SzB0bo6Dv/EWjFktE+P/les1PBxGmdHkD4yqy0vXbYWzLjMf+VRnvA323ZNzk4A4c6MsQC3PFuDgyP6LVlaJ5Sx7fIIGauydgDIzqePIdS/PLTJKIljV6tzkGzF2M9H4pvSq3nsgMrUKAR09jIbTLdaOxpcburtOqDlVwbHW7i7YngVNGNTvOGrQKtvL93VF+eTMVQUTQqoEGhdWn5JrNryfMdWFdC3vfA9r+aE7+r4SWWzdi/XaDWcXaSjGUveGSRJYUQrNvH2tKV+EGXYKNfXFxg6pK2hXmGsQZSyXwWVA7mzIsbKdkycRfGavWMYvrADWz/KDtwdg5LiuXkyUhgfwZrtAQE50k3Dlg3MLl27TQqc=~0~-1~-1; KP_UIDz-ssn=02rcgiwlC6QcdD4HSyDzO84ov5sraIHxL9qPlyrTOT4H6k7LyLD9RoKI4IgSwLDs7KAco9TPuQP45bKQSe1Eknb2gfgllUoQfVYhgsDzNpu8iPwnpY4FLfAoCrvQZY7vz74sJ1PpBV6vSDCbCkYkJbYa9PsRjlYQKwGb6Bd7sV; KP_UIDz=02rcgiwlC6QcdD4HSyDzO84ov5sraIHxL9qPlyrTOT4H6k7LyLD9RoKI4IgSwLDs7KAco9TPuQP45bKQSe1Eknb2gfgllUoQfVYhgsDzNpu8iPwnpY4FLfAoCrvQZY7vz74sJ1PpBV6vSDCbCkYkJbYa9PsRjlYQKwGb6Bd7sV; plpids=28fm:3qyi|284v:284w|2qlr:3dld|68i3:68i4|39ac:nvzy|286r:286s|2qnk:2ryt|2883:r8wm|3mr5:3mr6|do7i:gvye|2bgh:3tls|6e87:r93i|28h8:28h9|mbfy:q29q|2qlj:3dle|39ah:nvym|2s7y:r93a|gr7q:gr7y|3r12:3r13|k01q:k0gu|3l0l:3l0m|32g4:32g5|28bh:28bj|3mni:3mnj|3mr3:3mr4|28g9:r8xy|28a8:28a9|2ql8:2qla|lt8e:nhx2|28jz:28k0|iud2:iuda|l3ri:l3rq|4cuk:4cul|2qnc:2y4v|l3qm:l3qu|2855:r8yu; _ga_GM4GWYGVKP=GS1.1.1728372145.8.1.1728372164.41.0.0; _clsk=12df93h%7C1728372164793%7C3%7C0%7Cr.clarity.ms%2Fcollect; rxvt=1728373965004|1728372145522; akavpau_defaultvp=1728372465~id=b14ffcbfec292c291bbfdce4efcfabc0; akaalb_chewy_ALB=1728372765~op=prd_chewy_lando:www-prd-lando-use2|prd_chewy_plp:www-prd-plp-use2|prd_kasada_plp:prd-kasada-plp-use2|chewy_com_ALB:www-chewy-use2|kasada_prd:kasada-prd|prd_kasada:prd-kasada-haproxy-use1|~rv=86~m=www-prd-lando-use2:0|www-prd-plp-use2:0|prd-kasada-plp-use2:0|www-chewy-use2:0|kasada-prd:0|prd-kasada-haproxy-use1:0|~os=43a06daff4514d805d02d3b6b5e79808~id=e0256312fba6f14d298097936f1f798f; dtPC=12$372153273_297h-vCCMFKGHPEBLUOPKIIFDMIPHJKGPMSLIH-0e0; RT="z=1&dm=www.chewy.com&si=843cf466-6940-48be-95fd-c6abedbb3823&ss=m20457l9&sl=2&tt=4qz&rl=1&nu=3e3ez51&cl=igw&obo=1&ld=1efx&r=3e3ez51&ul=1efx"',
    'priority': 'u=0, i',
    'referer': 'https://www.chewy.com/',
    'sec-ch-ua': '"Google Chrome";v="129", "Not=A?Brand";v="8", "Chromium";v="129"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
}
    
def write_first_row(sh, first_row):
    column = 1
    for field in first_row:
        sh.cell(row=1, column=column).value = field
        sh.cell(row=1, column=column).font = Font(bold=True)
        column = column + 1
        
def get_links(sh, url):

    links = []
    for row in sh: 
        name = row[0].value.replace('https://www.chewy.com/', '').split('/')[0].replace('/', '')
        links.append(name)
    
    response = requests.get(url, cookies=cookies, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')
        
    items = soup.find_all("div", class_="kib-product-card")
    
    for item in items:
        link = item.find("a", class_="kib-product-title")["href"]
        
        link_stripped = link.replace('https://www.chewy.com/', '').split('/')[0]
        
        if link_stripped not in links:
        
            row = sh.max_row + 1
            
            if 'chewy.com' not in link:
                sh.cell(row=row, column=1).value = 'https://www.chewy.com' + link
            elif 'api/event' in link:
                sh.cell(row=row, column=1).value = link.split('redirect=')[1].strip()
            else:
                sh.cell(row=row, column=1).value = link
            
            links.append(link_stripped)

    time.sleep(1)
    
def get_details(sh_details, url, id_for_items):

    link_for_items = "https://www.chewy.com/_next/data/chewy-pdp-ui-"+id_for_items+"/en-US/"+url.replace("https://www.chewy.com/", "")+".json?id="+url.split('/')[-1]+"&slug="+url.split('/')[-3]
    
    response_items = requests.get(link_for_items, cookies=cookies, headers=headers)
    soup_items = BeautifulSoup(response_items.text, 'html.parser')
    
    if "__N_REDIRECT" in json.loads(soup_items.text)["pageProps"]:
        new_link = "https://www.chewy.com/_next/data/chewy-pdp-ui-"+id_for_items+"/en-US/" + json.loads(soup_items.text)["pageProps"]["__N_REDIRECT"] + ".json?id="+url.split('/')[-1]+"&slug="+url.split('/')[-3]
        response_items = requests.get(link_for_items, cookies=cookies, headers=headers)
        soup_items = BeautifulSoup(response_items.text, 'html.parser')
        
    items_json=json.loads(soup_items.text)["pageProps"]["__APOLLO_STATE__"]
    items = [key for key, val in items_json.items() if "Item" in key]
    
    products = [key for key, val in items_json.items() if "Product" in key]
    manufacturer = items_json[products[0]]["manufacturerName"]
    
    swatches_list = {}
    swatch_types = [val for key, val in items_json.items() if "MDA:" in key]
    for st in swatch_types:
        if "isEnsemble" in st:
            #print(st)
            new_swatch = []
            title = st["name"]
            options = st["options"]
            for option in options:
                swatch_variant = json.loads(('{' + option["__ref"] + '}').replace('AttributeValue', '"AttributeValue"'))["AttributeValue"]["value"].strip()
                new_swatch.append(swatch_variant)
            swatches_list[title] = new_swatch
            if title not in first_row:
                first_row.append(title)
                write_first_row(sh_details, first_row)
    
    categories = [val["name"] for key, val in items_json.items() if "Breadcrumb" in key]
    details_short = ''
    
    for item in items:
        if ("inStock" in items_json[item] and items_json[item]["inStock"] == True) or ("isInStock" in items_json[item] and items_json[item]["isInStock"] == True):            
            row = sh_details.max_row + 1            
            attributeValues = items_json[item]['attributeValues({"includeEnsemble":true,"usage":["DEFINING"]})']
            for attr in attributeValues:
                variants = json.loads(('{' + attr['__ref'] + '}').replace('AttributeValue', '"AttributeValue"'))
                variant = variants["AttributeValue"]["value"].strip()
                for swatch in swatches_list:
                    if variant in swatches_list[swatch]:
                        index = first_row.index(swatch)
                        sh_details.cell(row=row, column=index+1).value = variant
            
            details_long = items_json[item]["description"]
            details_short = ''
            if "keyBenefits" in items_json[item] and items_json[item]["keyBenefits"] != None:
                for benefit in items_json[item]["keyBenefits"]:
                    details_short = details_short + " " + benefit
            sh_details.cell(row=row, column=11).value = str(details_short) + " " + details_long
            sh_details.cell(row=row, column=2).value = items_json[item]["name"]
            sh_details.cell(row=row, column=3).value = manufacturer
            sh_details.cell(row=row, column=4).value = items_json[item]["advertisedPrice"]
            if "strikeThroughPrice" in items_json[item]:
                sh_details.cell(row=row, column=5).value = items_json[item]["strikeThroughPrice"]
            sh_details.cell(row=row, column=12).value = items_json[item]["partNumber"]
            sh_details.cell(row=row, column=1).value = url.replace(url.split('/')[-1], items_json[item]["entryID"])
            
            for category in range(len(categories)):       
                sh_details.cell(row=row, column=6+category).value = categories[category]
            
            images = []
            
            if "images" in items_json[item]:
                for image in items_json[item]["images"]:
                    images.append(image['url({"autoCrop":true,"sideLongest":1800})'])
            else:
                images.append(items_json[item]["fullImage"]['url({"autoCrop":true,"square":1800})'])
            
            
            for image in range(len(images)):
                sh_details.cell(row=row, column=13+image).value = images[image]
    
if 'stage' not in st.session_state:
    st.session_state.stage = 0

def set_stage(stage):
    st.session_state.stage = stage

workbook = openpyxl.Workbook()
sh = workbook.active
workbook_details = openpyxl.Workbook()
sh_details = workbook_details.active

first_row = ["URL", "Product Title", "Manufacturer", "Price", "Regular Price", "Category 1", "Category 2", "Category 3", "Category 4", "Category 5", "Details", "Item Number", "Photo1", "Photo2", "Photo3", "Photo4", "Photo5", "Photo6", "Photo7", "Photo8", "Photo9", "Photo10", "Photo11", "Photo12", "Photo13", "Photo14", "Photo15"]
write_first_row(sh, first_row)
write_first_row(sh_details, first_row)

st.write("Write your category here (format: dental-treats-1539):")
category = st.text_input("Category")

st.button('Go!', on_click=set_stage, args=(2,))

if st.session_state.stage > 0:
    
    if st.session_state.stage == 2:
        response = requests.get("https://www.chewy.com/b/" + category, cookies=cookies, headers=headers)
        if (response.status_code == 200):
            soup = BeautifulSoup(response.text, 'html.parser')
            max_page = 1
            max_page_container = soup.find_all('li', class_= 'kib-pagination-new__list-item')
            if max_page_container != []:
                max_page = max_page_container[-1].find('a')['aria-label'].replace("Page", "").strip()
            
            st.write("Scraping in progress...")
            
            st.write("Getting links..")
            for i in stqdm(range(int(max_page))):
                url = "https://www.chewy.com/b/" + category.rsplit("-", 1)[0] +"_c" + category.rsplit("-", 1)[1] + "_p" + str(i+1)
                get_links(sh, url)
                time.sleep(1)
            
            workbook.save("chewy.xlsx")

            id_for_items = ""

            st.write("Going through the products...")
            for row in stqdm (range(2, sh.max_row)):    
                url = sh[row+1][0].value

                if row == 2 or row % 50 == 1:
                    response = requests.get(url, cookies=cookies, headers=headers)
                    soup = BeautifulSoup(response.text, 'html.parser')    
                    scripts = soup.find("head").find_all("script")
                    for script in scripts:
                        #print(script)
                        if script.has_attr("src"):
                            if "/static/spa/chewy-pdp-ui/_next/static/chewy-pdp-ui-" in script["src"]:
                                id_for_items = script["src"].replace("/static/spa/chewy-pdp-ui/_next/static/chewy-pdp-ui-", "").replace("/_buildManifest.js", "").replace("/_ssgManifest.js", "")
                                break
                
                if "URL" not in url:
                    try:
                        get_details(sh_details, url, id_for_items)
                    except Exception as e:
                        print(e)
                time.sleep(1)
                
        
                workbook_details.save('chewy_data.xlsx')
            set_stage(1)

        elif (response.status_code == 404):
            st.write("Wrong category!")
        else:
            st.write("An error occured! Status code: " + str(response.status_code))
    
    st.write("Done!")

    with open("./chewy_data.xlsx", "rb") as file:
        st.download_button("Download file", data=file, file_name=("chewy-"+category+".xlsx"), mime="application/vnd.ms-excel")
    
    
